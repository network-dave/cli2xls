#!/usr/bin/env python3

'''

Name:           json2xls.py
Description:    Read JSON data parsed with ntc-templates and write a table in XLS or CSV format
Author:         David Paneels

Usage:          see json2xls.py --help

'''

import argparse
import sys
import os
import json
import csv
import re
import logging

from openpyxl import Workbook
from openpyxl import load_workbook

import cli2json


# Global options
DELIMITER = ";"


def parse_json_to_table(json_data):
    '''
    Convert ntc-templates JSON data to a list array
    '''
    # Get all keys from JSON items and build column header
    table = []
    header = []
    for item in json_data:
        for key, _ in item.items():
            if key not in header:
                header.append(key)

    # Get value for each key in header (if key does not exist returns "N/A")
    for item in json_data:
        row = []
        for key in header:
            row.append(item.get(key, "N/A"))
        table.append(row)

    # Make header uppercase and prepend row to table
    header = [key.upper() for key in header]
    table = [header] + table
    return table

def add_table_to_workbook(table, filename, sheetname="default"):
    '''
    Add table data to a new worksheet in an Excel file
    '''
    # If XLS file exists, insert the table in a new worksheet else create a new workbook file
    if os.path.exists(filename):
        wb = load_workbook(filename)
        logging.info(f"[+] Workbook {filename} already exists, adding new worksheet to it")
        if sheetname in wb.sheetnames:
            ws = wb.create_sheet(sheetname + " NEW")
        else:
            ws = wb.create_sheet(sheetname)
    else:
        logging.info(f"[+] Creating new Excel workbook {filename}")
        wb = Workbook()
        ws = wb.active
        ws.title = sheetname

    for row in table:
        # Convert each item in the row in a string to avoid type errors when appending
        ws.append([str(cell) for cell in row])
    
    wb.save(filename = filename)


def main():
    '''
    Main program
    '''
    # Parse command line arguments
    argparser = argparse.ArgumentParser(
        description="Read JSON data and write an XLS or CSV file"
        )
    argparser.add_argument(
        "-i",
        "--infile",
        metavar="filename.json",
        required=True,
        type=argparse.FileType("r"),
        help="JSON file (use '-' to read from stdin)",
        )
    argparser.add_argument(
        "-o",
        "--outfile",
        metavar="filename.xls",
        help="save output to Excel or CSV file (supported formats: .xls, .xlsx, .csv)"
        )
    argparser.add_argument(
        "--verbose",
        action="store_true",
        help="print additional information to stderr"
        )
    args = argparser.parse_args()

    # Configure logging
    if args.verbose:
        logging.basicConfig(format="%(message)s", level=logging.INFO)
    else:
        logging.basicConfig(format="%(message)s", level=logging.WARNING)

    # Load and deserialize JSON data from filehandler (it has already been opened by argparse)
    with args.infile as f:
        infilename = f.name
        logging.info(f"[+] Loading JSON data from {infilename}")
        json_data = json.loads(f.read())

    # If no data is present we'll just exit with error code 1
    if not json_data:
        logging.warning("[!] JSON data structure is empty")
        sys.exit(1)

    table = parse_json_to_table(json_data)

    # Export the data according to the file type specified
    if args.outfile:
        if ".csv" in args.outfile:
            with open(args.outfile, "w") as f:
                c = csv.writer(f, delimiter=DELIMITER)
                for row in table:
                    c.writerow(row)
        elif ".xls" in args.outfile:        # Also works for .xlsx
            add_table_to_workbook(
                table, 
                args.outfile, 
                sheetname=cli2json.get_parser_from_filename(infilename)
                )
        logging.info(f"[+] Done writing data to {args.outfile}")
    else:
        # If no output file is specified just print raw text to stdout
        for row in table:
            print("\t".join(str(s) for s in row))


if __name__ == "__main__":
    '''
    main()
    '''
    try:
        main()
    except KeyboardInterrupt:
        print()
        sys.exit(1)
    except json.decoder.JSONDecodeError:
        logging.critical("[!] An error occured while decoding the JSON data (exit code 1)")
        sys.exit(1)
    except Exception as e:
        logging.critical(f"[!] An error occured during the operation ({str(e)})")
        sys.exit(1)
    